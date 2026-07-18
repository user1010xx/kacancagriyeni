from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from invekto_client import _call_datetime, _department_name, parse_call_datetime


def sort_calls(calls: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Proper datetime sorting (fixes lexical dd.mm.yyyy bug across months)."""
    def sort_key(call: dict[str, Any]):
        parsed = parse_call_datetime(call)
        if parsed:
            return parsed
        # Fallback to string
        d, t = _call_datetime(call)
        return (d, t)
    return sorted(calls, key=sort_key)


def _get_call_field(call: dict[str, Any], *keys: str, default: str = "") -> str:
    for k in keys:
        val = call.get(k)
        if val not in (None, ""):
            return str(val).strip()
    return default


def export_missed_calls_excel(
    calls: list[dict[str, Any]],
    output_path: Path,
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Kaçan Çağrılar"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # Headers - rich data
    headers = [
        "ID", "Telefon", "Tarih", "Saat", "Departman/Kuyruk",
        "Durum (Status)", "Tamamlandı (IsCompleted)", "Çağrı Süresi",
        "Ring Süresi", "Bekleme Süresi", "Trunk", "Extension"
    ]

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, call in enumerate(calls, start=2):
        call_date, call_time = _call_datetime(call)
        dept = _department_name(call)

        # Extract rich fields (support both report types)
        call_id = _get_call_field(call, "ID", "CallID", default="")
        phone = _get_call_field(call, "Phone", default="Bilinmiyor")
        status = _get_call_field(call, "Status", default="")
        is_completed = str(call.get("IsCompleted", "")).strip()
        call_time_val = _get_call_field(call, "CallTime", "CallTimeSecond", default="")
        ring_time = _get_call_field(call, "RingTime", "RingDuration", "RingTimeSecond", default="")
        wait_time = _get_call_field(call, "WaitTime", default="")
        trunk = _get_call_field(call, "Trunk", default="")
        extension = _get_call_field(call, "Extension", "ExtensionName", "CompletedExtension", "CompletedExtensionName", default="")

        values = [
            call_id,
            phone,
            call_date,
            call_time,
            dept,
            status,
            is_completed,
            call_time_val,
            ring_time,
            wait_time,
            trunk,
            extension,
        ]

        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col in (3, 4):  # date/time center
                cell.alignment = center_align

    # Column widths (using proper openpyxl helper)
    widths = [12, 16, 12, 10, 22, 14, 18, 14, 12, 14, 10, 18]
    for i, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    sheet.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def export_delivered_report_excel(
    rows: list[dict],
    output_path: Path,
) -> Path:
    """Personele başarıyla iletilen kaçan çağrılar: Numara, Personel Adı, İletilen Saat, Geri Arama Durumu."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "İletilen Çağrılar"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    headers = ["Personel Adı", "Numara", "İletilen Saat", "Geri Arama Durumu"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, row in enumerate(rows, start=2):
        values = [
            row.get("personel_adi", ""),
            row.get("phone", ""),
            row.get("notified_at", ""),
            row.get("callback_status", ""),
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col == 3:
                cell.alignment = center_align

    widths = [24, 18, 22, 26]
    for i, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = w

    sheet.freeze_panes = "A2"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path