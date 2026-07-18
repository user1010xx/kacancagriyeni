import json
import os
import threading
from pathlib import Path

from openpyxl import load_workbook


class PersonnelStore:
    """Personel yönetimi: dahili_ad -> {personel_adi, telegram_username, telegram_chat_id}"""

    def __init__(self, path: Path) -> None:
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._lock = threading.Lock()
        self._data: dict[str, dict[str, str]] = {}
        self._load()

    def _load(self) -> None:
        if not self.path.exists():
            return
        try:
            with self.path.open("r", encoding="utf-8") as f:
                loaded = json.load(f)
            if isinstance(loaded, dict):
                self._data = loaded
        except Exception:
            self._data = {}

    def _save(self) -> None:
        temp_path = self.path.with_suffix(".tmp")
        with temp_path.open("w", encoding="utf-8") as f:
            json.dump(self._data, f, ensure_ascii=False, indent=2)
        os.replace(temp_path, self.path)

    def add_or_update(
        self,
        dahili_ad: str,
        personel_adi: str,
        telegram_username: str,
        *,
        telegram_chat_id: str | None = None,
        save: bool = True,
    ) -> bool:
        dahili = str(dahili_ad).strip()
        if not dahili:
            return False

        existing = self._data.get(dahili, {})
        chat_id = (
            str(telegram_chat_id).strip()
            if telegram_chat_id is not None
            else existing.get("telegram_chat_id", "")
        )

        self._data[dahili] = {
            "personel_adi": str(personel_adi).strip(),
            "telegram_username": str(telegram_username).strip().lstrip("@"),
            "telegram_chat_id": chat_id,
        }
        if save:
            with self._lock:
                self._save()
        return True

    def link_chat_id_by_username(self, username: str, chat_id: int) -> int:
        if not username:
            return 0

        normalized = str(username).strip().lstrip("@").casefold()
        updated = 0
        with self._lock:
            for dahili, info in self._data.items():
                stored = str(info.get("telegram_username", "")).strip().lstrip("@").casefold()
                if stored and stored == normalized:
                    info["telegram_chat_id"] = str(chat_id)
                    updated += 1
            if updated:
                self._save()
        return updated

    def remove(self, dahili_ad: str) -> bool:
        dahili = str(dahili_ad).strip()
        with self._lock:
            if dahili in self._data:
                del self._data[dahili]
                self._save()
                return True
        return False

    @staticmethod
    def _extension_token(value: str) -> str:
        """Invekto/UI farklarını yumuşatır: 'selen-K' ve 'Selen K' -> 'selen'."""
        text = str(value).strip().casefold()
        if not text:
            return ""
        return text.split("-")[0].split()[0]

    def get(self, dahili_ad: str) -> dict[str, str] | None:
        return self._data.get(str(dahili_ad).strip())

    def find_for_extension(self, extension: str) -> dict[str, str] | None:
        """Invekto ExtensionName ile personel kaydını eşleştirir.

        Sıra: tam dahili anahtarı -> büyük/küçük harf -> personel adı -> @username
        """
        ext = str(extension).strip()
        if not ext:
            return None

        direct = self.get(ext)
        if direct:
            return direct

        ext_cf = ext.casefold()
        ext_token = self._extension_token(ext)

        for dahili, info in self._data.items():
            if str(dahili).strip().casefold() == ext_cf:
                return info

        for info in self._data.values():
            ad = str(info.get("personel_adi", "")).strip()
            if not ad:
                continue
            if ad.casefold() == ext_cf or self._extension_token(ad) == ext_token:
                return info

            username = str(info.get("telegram_username", "")).strip().lstrip("@")
            if username and (
                username.casefold() == ext_cf
                or self._extension_token(username) == ext_token
            ):
                return info

        return None

    def get_all(self) -> list[dict[str, str]]:
        result = []
        for dahili, info in self._data.items():
            chat_id = str(info.get("telegram_chat_id", "")).strip()
            result.append(
                {
                    "dahili_ad": dahili,
                    "personel_adi": info.get("personel_adi", ""),
                    "telegram_username": info.get("telegram_username", ""),
                    "telegram_chat_id": chat_id,
                    "dm_ready": bool(chat_id),
                }
            )
        return sorted(result, key=lambda x: x["dahili_ad"])

    def load_from_excel(self, excel_path: Path) -> int:
        if not excel_path.exists():
            return 0

        count = 0
        wb = load_workbook(excel_path, read_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                dahili = str(row[0]).strip() if row[0] else ""
                ad = str(row[1]).strip() if row[1] else ""
                username = str(row[2]).strip() if len(row) > 2 and row[2] else ""

                if not dahili or dahili.lower() in {
                    "dahili",
                    "dahili_ad",
                    "extension",
                    "extensionname",
                }:
                    continue
                if not ad:
                    continue

                if self.add_or_update(dahili, ad, username, save=False):
                    count += 1
        finally:
            wb.close()

        if count:
            with self._lock:
                self._save()
        return count

    def count(self) -> int:
        return len(self._data)